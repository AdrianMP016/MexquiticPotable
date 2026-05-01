#!/usr/bin/env python3
import json
import re
import sys
import unicodedata
from collections import OrderedDict

from openpyxl import load_workbook


REQUIRED = {"id", "usuario", "medidor", "ruta", "colonia"}
ALIASES = {
    "direccion": {"direccion", "dirección", "domicilio"},
    "observaciones": {"observaciones", "observacion", "obs"},
}


def normalize_header(value):
    text = "" if value is None else str(value).strip().lower()
    text = "".join(
        char for char in unicodedata.normalize("NFD", text)
        if unicodedata.category(char) != "Mn"
    )
    text = re.sub(r"\s+", " ", text)
    return text


def clean_value(value):
    if value is None:
        return None
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text or None


def build_index(headers):
    index = {}
    for idx, header in enumerate(headers):
        if header and header not in index:
            index[header] = idx
    return index


def pick_index(index, *names):
    for name in names:
        if name in index:
            return index[name]
    return None


def row_key(record):
    if record.get("id_excel"):
        return f"ID:{record['id_excel']}"
    if record.get("medidor"):
        medidor = re.sub(r"[^A-Z0-9-]", "", record["medidor"].upper())
        return f"MEDIDOR:{medidor}"
    return None


def get_valid_sheet_records(sheet):
    rows = sheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        return None

    headers = [normalize_header(value) for value in header_row]
    index = build_index(headers)

    has_required = REQUIRED.issubset(set(headers))
    direccion_idx = pick_index(index, *ALIASES["direccion"])
    colonia_idx = pick_index(index, "colonia")
    observaciones_idx = pick_index(index, *ALIASES["observaciones"])

    if not has_required or direccion_idx is None or colonia_idx is None:
        return None

    merged = OrderedDict()

    for row_number, row in enumerate(rows, start=2):
        if not row or all(value is None or str(value).strip() == "" for value in row):
            continue

        record = {
            "id_excel": clean_value(row[index["id"]]) if index["id"] < len(row) else None,
            "usuario": clean_value(row[index["usuario"]]) if index["usuario"] < len(row) else None,
            "medidor": clean_value(row[index["medidor"]]) if index["medidor"] < len(row) else None,
            "ruta": clean_value(row[index["ruta"]]) if index["ruta"] < len(row) else None,
            "domicilio": clean_value(row[direccion_idx]) if direccion_idx < len(row) else None,
            "colonia": clean_value(row[colonia_idx]) if colonia_idx < len(row) else None,
            "observaciones": clean_value(row[observaciones_idx]) if observaciones_idx is not None and observaciones_idx < len(row) else None,
            "source_sheet": sheet.title,
            "source_row": row_number,
        }

        key = row_key(record)
        if not key:
            continue

        merged[key] = record

    return list(merged.values())


def pick_preferred_sheet(workbook):
    valid = []

    for sheet in workbook.worksheets:
        records = get_valid_sheet_records(sheet)
        if records is None:
            continue
        valid.append((sheet, records))

    if not valid:
        return None, None

    for sheet, records in valid:
        if normalize_header(sheet.title) == "hoja2":
            return sheet, records

    return valid[-1]


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Debes indicar la ruta del Excel."}))
        sys.exit(1)

    path = sys.argv[1]

    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        print(json.dumps({"error": f"No se pudo abrir el Excel: {exc}"}))
        sys.exit(1)

    selected_sheet, records = pick_preferred_sheet(workbook)

    if selected_sheet is None or records is None:
        print(json.dumps({"error": "No se encontro ninguna hoja valida con las columnas requeridas."}))
        sys.exit(1)

    payload = {
        "rows": records,
        "meta": {
            "used_sheets": [selected_sheet.title],
            "row_count": len(records),
        },
    }
    print(json.dumps(payload))


if __name__ == "__main__":
    main()
